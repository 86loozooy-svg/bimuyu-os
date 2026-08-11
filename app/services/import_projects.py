"""P2 离线模板闭环：模板生成 + 导入解析。

设计约束（与 P1 一致，保证内网离线可用）：
- 完全基于 openpyxl（后端 Python）生成与解析，不引入 Node/前端 CDN 库。
- 前端只负责上传文件、渲染后端返回的预览 JSON、提交确认，无第三方 JS 依赖。

模板列（顺序即表头）：
  A 项目名称*  B 客户  C 行业  D 面积(㎡)  E 工种(下拉)  F 数量
  G 单价(¥)    H 金额(¥)=F*G  I 税率(%)  J 税额(¥)=H*I/100  K 含税合计(¥)=H+J  L 状态(下拉)
"""
from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.database import db_session, get_connection, rows_to_list

# 模板列（名称, 是否必填）
TEMPLATE_HEADERS = [
    ("项目名称", True),
    ("客户", False),
    ("行业", False),
    ("面积(㎡)", False),
    ("工种", False),
    ("数量", False),
    ("单价(¥)", False),
    ("金额(¥)", False),
    ("税率(%)", False),
    ("税额(¥)", False),
    ("含税合计(¥)", False),
    ("状态", False),
]

STATUS_OPTIONS = ["lead", "quote", "design", "construction", "done"]
HEADER_FILL = "3B332A"  # 复用 BOM 导出暗色表头
MONEY_FMT = "#,##0.00"
ERROR_FILL = "FFC7CE"  # 条件格式：异常行标红


def _labor_names() -> list:
    with db_session() as conn:
        rows = rows_to_list(
            conn.execute("SELECT name FROM labor_items ORDER BY name").fetchall()
        )
    return [r["name"] for r in rows]


def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num(v):
    """返回 (value, valid)。空=合法(None)，非数字=invalid。"""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return None, True
    try:
        return float(v), True
    except (ValueError, TypeError):
        return None, False


# ── 模板生成 ────────────────────────────────────────────────────────────────
def build_template_bytes(studio: dict) -> bytes:
    labor = _labor_names()
    wb = Workbook()
    ws = wb.active
    ws.title = "项目导入"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")

    # 标题行
    ws.merge_cells("A1:L1")
    ws["A1"] = f"{studio.get('name', 'Studio OS')} · 项目批量导入模板"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    # 表头（第 2 行）
    for col, (h, _req) in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 预置公式行（第 3 行起，共 200 行）
    first, last = 3, 202
    for r in range(first, last + 1):
        ws.cell(row=r, column=8, value=f'=IF(OR(F{r}="",G{r}=""),"",F{r}*G{r})')   # H 金额
        ws.cell(row=r, column=9, value=0)                                            # I 税率默认 0
        ws.cell(row=r, column=10, value=f'=IF(OR(H{r}="",I{r}=""),"",H{r}*I{r}/100)')  # J 税额
        ws.cell(row=r, column=11, value=f'=IF(OR(H{r}="",J{r}=""),"",H{r}+J{r})')   # K 含税合计
        for col in (6, 7, 8, 10, 11):
            ws.cell(row=r, column=col).number_format = MONEY_FMT

    # 列宽
    widths = [24, 16, 12, 10, 14, 8, 12, 12, 10, 12, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 下拉：工种（引字典）+ 状态
    if labor:
        dv_labor = DataValidation(
            type="list", formula1=f'"{",".join(labor)}"', allow_blank=True
        )
        dv_labor.add(f"E{first}:E{last}")
        ws.add_data_validation(dv_labor)
    dv_status = DataValidation(
        type="list", formula1=f'"{",".join(STATUS_OPTIONS)}"', allow_blank=True
    )
    dv_status.add(f"L{first}:L{last}")
    ws.add_data_validation(dv_status)

    # 条件格式：项目名称为空 / 金额为负 → 红底
    red = PatternFill(start_color=ERROR_FILL, end_color=ERROR_FILL, fill_type="solid")
    ws.conditional_formatting.add(
        f"A{first}:A{last}", CellIsRule(operator="equal", formula=['""'], fill=red)
    )
    ws.conditional_formatting.add(
        f"H{first}:H{last}", CellIsRule(operator="lessThan", formula=["0"], fill=red)
    )

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── 导入解析 ────────────────────────────────────────────────────────────────
def parse_template_bytes(content: bytes) -> dict:
    """解析上传 xlsx，返回 {rows, summary}。公式单元格读不到计算值，金额本地重算。"""
    labor = set(_labor_names())
    wb = load_workbook(io.BytesIO(content), data_only=False)
    ws = wb.active
    rows_out: list = []
    invalid = 0
    total = 0

    max_r = min(ws.max_row, 1000)
    r = 3
    while r <= max_r:
        row_cells = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        if all(v in (None, "") for v in row_cells):
            r += 1
            continue

        name = _str(row_cells[0])
        client = _str(row_cells[1])
        industry = _str(row_cells[2])
        area, area_ok = _num(row_cells[3])
        labor_name = _str(row_cells[4])
        qty, qty_ok = _num(row_cells[5])
        price, price_ok = _num(row_cells[6])
        tax_rate, tax_ok = _num(row_cells[8])
        status = _str(row_cells[11]) or "lead"

        # 金额本地重算（不依赖公式单元格）
        amount = round(qty * price, 2) if (qty is not None and price is not None) else None
        tax_rate = tax_rate or 0.0
        tax = round(amount * tax_rate / 100, 2) if amount is not None else None
        total_amt = round(amount + tax, 2) if amount is not None else None

        errors = []
        if not name:
            errors.append("项目名称为必填")
        if not area_ok:
            errors.append("面积不是有效数字")
        if area is not None and area < 0:
            errors.append("面积不能为负")
        if not qty_ok:
            errors.append("数量不是有效数字")
        if qty is not None and qty < 0:
            errors.append("数量不能为负")
        if not price_ok:
            errors.append("单价不是有效数字")
        if price is not None and price < 0:
            errors.append("单价不能为负")
        if not tax_ok:
            errors.append("税率不是有效数字")
        if labor_name and labor_name not in labor:
            errors.append("工种不在字典内")
        if status not in STATUS_OPTIONS:
            errors.append("状态非法")

        ok = len(errors) == 0
        if not ok:
            invalid += 1
        total += 1
        rows_out.append({
            "row": r,
            "name": name,
            "client": client,
            "industry": industry,
            "area": area,
            "labor": labor_name,
            "qty": qty,
            "price": price,
            "amount": amount,
            "tax_rate": tax_rate,
            "tax": tax,
            "total": total_amt,
            "status": status,
            "ok": ok,
            "errors": errors,
        })
        r += 1

    return {
        "rows": rows_out,
        "summary": {"total": total, "valid": total - invalid, "invalid": invalid},
    }
