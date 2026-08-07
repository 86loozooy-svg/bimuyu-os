"""第一批功能：项目记账 & 预算追踪 + 物料清单 BOM（CRUD + 导出）。

路由前缀 /api/projects：
- GET  /{pid}/budget                    预算聚合（卡片 + 折线图数据）
- POST /{pid}/budget-items              新增分项
- PUT  /{pid}/budget-items/{iid}        改分项
- DELETE /{pid}/budget-items/{iid}      删分项（关联开销置空 budget_item_id）
- GET  /{pid}/expenses                  开销列表
- POST /{pid}/expenses                  记一笔开销（含附件）
- DELETE /{pid}/expenses/{eid}          删开销
- GET  /{pid}/materials                 BOM 列表
- POST /{pid}/materials                 新增物料
- PUT  /{pid}/materials/{mid}           改物料
- DELETE /{pid}/materials/{mid}         删物料
- POST /{pid}/materials/generate-from-cost  从造价生成草稿
- GET  /{pid}/materials/export/excel
- GET  /{pid}/materials/export/pdf
"""

import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, Response

from app.auth import get_current_user, user_can_access_project
from app.config import BASE_DIR
from app.database import db_session, get_connection, row_to_dict, rows_to_list

router = APIRouter(prefix="/api/projects", tags=["budget"])

EXPENSE_UPLOAD_DIR = BASE_DIR / "data" / "uploads" / "expenses"
EXPENSE_MAX_BYTES = 10 * 1024 * 1024
_ALLOWED_ATTACH_EXT = {
    b"%PDF": ".pdf",
    b"\xff\xd8\xff": ".jpg",
    b"\x89PNG\r\n\x1a\n": ".png",
    b"RIFF": ".webp",
}


def _require_editor(user: dict) -> None:
    if user.get("role") == "viewer":
        raise HTTPException(status_code=403, detail="访客无编辑权限")


def _detect_ext(content: bytes) -> str | None:
    for magic, ext in _ALLOWED_ATTACH_EXT.items():
        if content.startswith(magic):
            if ext == ".webp":
                if b"WEBP" in content[:16]:
                    return ".webp"
                continue
            return ext
    return None


async def save_expense_attachment(upload: UploadFile) -> str:
    content = await upload.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="空文件")
    if len(content) > EXPENSE_MAX_BYTES:
        raise HTTPException(status_code=413, detail="附件过大（≤10MB）")
    ext = _detect_ext(content)
    if ext is None:
        raise HTTPException(status_code=400, detail="仅支持 PDF / JPG / PNG / WebP")
    filename = f"{uuid.uuid4().hex}{ext}"
    EXPENSE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    (EXPENSE_UPLOAD_DIR / filename).write_bytes(content)
    return f"/media/uploads/expenses/{filename}"


# ── 预算聚合 ──────────────────────────────────────────────────────────────
def _aggregate_budget(conn, pid: int) -> dict:
    items = rows_to_list(
        conn.execute(
            "SELECT * FROM project_budget_item WHERE project_id=? ORDER BY sort_order, id",
            (pid,),
        ).fetchall()
    )
    expenses = rows_to_list(
        conn.execute("SELECT * FROM project_expense WHERE project_id=?", (pid,)).fetchall()
    )
    spent_by_item: dict[int, float] = {}
    total_spent = 0.0
    for e in expenses:
        amt = float(e.get("amount") or 0)
        total_spent += amt
        bi = e.get("budget_item_id")
        if bi:
            spent_by_item[bi] = spent_by_item.get(bi, 0.0) + amt

    out_items = []
    labels, cum_budget, cum_actual = [], [], []
    run_b = run_a = 0.0
    for it in items:
        iid = it["id"]
        planned = float(it.get("planned_amount") or 0)
        spent = spent_by_item.get(iid, 0.0)
        remaining = round(planned - spent, 2)
        pct = round(spent / planned * 100, 1) if planned else 0.0
        out_items.append(
            {
                "id": iid,
                "name": it.get("name"),
                "category": it.get("category") or "其他",
                "planned": planned,
                "spent": round(spent, 2),
                "remaining": remaining,
                "pct": pct,
                "over": remaining < 0,
            }
        )
        run_b += planned
        run_a += spent
        labels.append(it.get("name") or f"#{iid}")
        cum_budget.append(round(run_b))
        cum_actual.append(round(run_a))

    total_planned = round(sum(float(i.get("planned_amount") or 0) for i in items), 2)
    return {
        "items": out_items,
        "total_planned": total_planned,
        "total_spent": round(total_spent, 2),
        "total_remaining": round(total_planned - total_spent, 2),
        "chart": {"labels": labels, "cum_budget": cum_budget, "cum_actual": cum_actual},
    }


@router.get("/{pid}/budget")
async def get_budget(pid: int, user: dict = Depends(get_current_user)):
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403, detail="无权访问")
    with db_session() as conn:
        data = _aggregate_budget(conn, pid)
    return JSONResponse(data)


# ── 预算分项 ──────────────────────────────────────────────────────────────
@router.post("/{pid}/budget-items")
async def add_budget_item(pid: int, request: Request, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    payload = await request.json()
    name = (payload.get("name") or "").strip()
    if not name:
        return JSONResponse({"error": "分项名称必填"}, status_code=400)
    planned = float(payload.get("planned_amount") or 0)
    category = (payload.get("category") or "其他").strip() or "其他"
    with db_session() as conn:
        cur = conn.execute(
            "INSERT INTO project_budget_item (project_id, name, category, planned_amount) VALUES (?,?,?,?)",
            (pid, name, category, planned),
        )
        iid = cur.lastrowid
        item = row_to_dict(
            conn.execute("SELECT * FROM project_budget_item WHERE id=?", (iid,)).fetchone()
        )
    return JSONResponse({"ok": True, "item": dict(item)})


@router.put("/{pid}/budget-items/{iid}")
async def update_budget_item(
    pid: int, iid: int, request: Request, user: dict = Depends(get_current_user)
):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    payload = await request.json()
    with db_session() as conn:
        row = conn.execute(
            "SELECT * FROM project_budget_item WHERE id=? AND project_id=?", (iid, pid)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404)
        name = (payload.get("name") or "").strip() or row["name"]
        category = (payload.get("category") or "").strip() or row["category"] or "其他"
        planned = float(payload.get("planned_amount", row["planned_amount"]) or 0)
        conn.execute(
            "UPDATE project_budget_item SET name=?, category=?, planned_amount=? WHERE id=?",
            (name, category, planned, iid),
        )
    return JSONResponse({"ok": True})


@router.delete("/{pid}/budget-items/{iid}")
async def delete_budget_item(pid: int, iid: int, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    with db_session() as conn:
        conn.execute("UPDATE project_expense SET budget_item_id=NULL WHERE budget_item_id=?", (iid,))
        conn.execute(
            "DELETE FROM project_budget_item WHERE id=? AND project_id=?", (iid, pid)
        )
    return JSONResponse({"ok": True})


# ── 开销流水 ──────────────────────────────────────────────────────────────
@router.get("/{pid}/expenses")
async def list_expenses(pid: int, user: dict = Depends(get_current_user)):
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    with db_session() as conn:
        rows = rows_to_list(
            conn.execute(
                "SELECT * FROM project_expense WHERE project_id=? ORDER BY occurred_date DESC, id DESC",
                (pid,),
            ).fetchall()
        )
    return JSONResponse({"expenses": rows})


@router.post("/{pid}/expenses")
async def add_expense(
    pid: int,
    request: Request,
    user: dict = Depends(get_current_user),
    budget_item_id: int = Form(None),
    amount: float = Form(...),
    payee: str = Form(""),
    occurred_date: str = Form(""),
    note: str = Form(""),
    attachment: UploadFile = File(None),
):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return JSONResponse({"error": "金额必须为数字"}, status_code=400)
    if amount <= 0:
        return JSONResponse({"error": "金额必须为正数"}, status_code=400)
    att_path = ""
    if attachment and attachment.filename:
        att_path = await save_expense_attachment(attachment)
    with db_session() as conn:
        conn.execute(
            "INSERT INTO project_expense "
            "(project_id, budget_item_id, amount, payee, occurred_date, note, attachment_path, created_by) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (pid, budget_item_id, amount, payee, occurred_date, note, att_path, user.get("id")),
        )
    return JSONResponse({"ok": True, "attachment": att_path})


@router.delete("/{pid}/expenses/{eid}")
async def delete_expense(pid: int, eid: int, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    with db_session() as conn:
        conn.execute("DELETE FROM project_expense WHERE id=? AND project_id=?", (eid, pid))
    return JSONResponse({"ok": True})


# ── 物料清单 BOM ──────────────────────────────────────────────────────────
@router.get("/{pid}/materials")
async def list_materials(pid: int, user: dict = Depends(get_current_user)):
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    with db_session() as conn:
        rows = rows_to_list(
            conn.execute(
                "SELECT * FROM project_material WHERE project_id=? ORDER BY sort_order, id",
                (pid,),
            ).fetchall()
        )
    return JSONResponse({"materials": rows})


@router.post("/{pid}/materials")
async def add_material(pid: int, request: Request, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    p = await request.json()
    name = (p.get("name") or "").strip()
    if not name:
        return JSONResponse({"error": "物料名称必填"}, status_code=400)
    with db_session() as conn:
        cur = conn.execute(
            "INSERT INTO project_material "
            "(project_id, name, brand, spec, quantity, unit, unit_price, category, purchase_stage, status) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                pid,
                name,
                p.get("brand", ""),
                p.get("spec", ""),
                float(p.get("quantity") or 1),
                p.get("unit", ""),
                float(p.get("unit_price") or 0),
                p.get("category") or "其他",
                p.get("purchase_stage", ""),
                p.get("status") or "pending",
            ),
        )
        mid = cur.lastrowid
        row = row_to_dict(conn.execute("SELECT * FROM project_material WHERE id=?", (mid,)).fetchone())
    return JSONResponse({"ok": True, "material": dict(row)})


@router.put("/{pid}/materials/{mid}")
async def update_material(
    pid: int, mid: int, request: Request, user: dict = Depends(get_current_user)
):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    p = await request.json()
    with db_session() as conn:
        row = conn.execute(
            "SELECT * FROM project_material WHERE id=? AND project_id=?", (mid, pid)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404)
        conn.execute(
            "UPDATE project_material SET name=?, brand=?, spec=?, quantity=?, unit=?, "
            "unit_price=?, category=?, purchase_stage=?, status=? WHERE id=?",
            (
                p.get("name", row["name"]),
                p.get("brand", row["brand"]),
                p.get("spec", row["spec"]),
                float(p.get("quantity", row["quantity"]) or 1),
                p.get("unit", row["unit"]),
                float(p.get("unit_price", row["unit_price"]) or 0),
                p.get("category", row["category"]) or "其他",
                p.get("purchase_stage", row["purchase_stage"]),
                p.get("status", row["status"]) or "pending",
                mid,
            ),
        )
    return JSONResponse({"ok": True})


@router.delete("/{pid}/materials/{mid}")
async def delete_material(pid: int, mid: int, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    with db_session() as conn:
        conn.execute("DELETE FROM project_material WHERE id=? AND project_id=?", (mid, pid))
    return JSONResponse({"ok": True})


@router.post("/{pid}/materials/generate-from-cost")
async def generate_from_cost(pid: int, request: Request, user: dict = Depends(get_current_user)):
    _require_editor(user)
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    from app.routers.cost_estimate import load_cost_base

    base = load_cost_base()
    non_material = {
        "version", "meta", "categories", "city_tiers", "old_renovation",
        "floor_height", "zone_complexity", "service_scope",
    }
    cats = [k for k in base.keys() if k not in non_material and isinstance(base[k], dict)]
    created = []
    with db_session() as conn:
        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM project_material WHERE project_id=?", (pid,)
        ).fetchone()[0]
        sort = max_sort
        for cat_key in cats:
            for mat_name, per_sqm in base[cat_key].items():
                sort += 1
                conn.execute(
                    "INSERT INTO project_material "
                    "(project_id, name, brand, spec, quantity, unit, unit_price, category, purchase_stage, status, sort_order) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        pid, mat_name, "", f"参考 +¥{per_sqm}/㎡", 1, "", 0,
                        cat_key, "", "pending", sort,
                    ),
                )
                created.append(mat_name)
    return JSONResponse({"ok": True, "count": len(created), "created": created})


# ── 导出 ──────────────────────────────────────────────────────────────────
@router.get("/{pid}/materials/export/excel")
async def export_materials_excel(pid: int, user: dict = Depends(get_current_user)):
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    conn = get_connection()
    try:
        project = row_to_dict(conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone())
        studio = row_to_dict(conn.execute("SELECT * FROM studio_profile WHERE id=1").fetchone())
        materials = rows_to_list(
            conn.execute(
                "SELECT * FROM project_material WHERE project_id=? ORDER BY category, sort_order, id",
                (pid,),
            ).fetchall()
        )
    finally:
        conn.close()
    data = export_bom_excel(materials, project or {}, studio or {})
    code = (project or {}).get("code", pid)
    return Response(
        content=bytes(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="BOM-{code}.xlsx"'},
    )


@router.get("/{pid}/materials/export/pdf")
async def export_materials_pdf(pid: int, user: dict = Depends(get_current_user)):
    if not user_can_access_project(user, pid):
        raise HTTPException(status_code=403)
    conn = get_connection()
    try:
        project = row_to_dict(conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone())
        studio = row_to_dict(conn.execute("SELECT * FROM studio_profile WHERE id=1").fetchone())
        materials = rows_to_list(
            conn.execute(
                "SELECT * FROM project_material WHERE project_id=? ORDER BY category, sort_order, id",
                (pid,),
            ).fetchall()
        )
    finally:
        conn.close()
    data = export_bom_pdf(materials, project or {}, studio or {})
    code = (project or {}).get("code", pid)
    return Response(
        content=bytes(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="BOM-{code}.pdf"'},
    )


# ── 导出实现 ──────────────────────────────────────────────────────────────
def export_bom_excel(materials: list, project: dict, studio: dict) -> bytes:
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cats = []
    for m in materials:
        c = m.get("category") or "其他"
        if c not in cats:
            cats.append(c)
    if not cats:
        cats = ["物料"]

    wb = Workbook()
    ws = wb.active
    ws.title = cats[0][:31]
    for cat in cats[1:]:
        wb.create_sheet(title=cat[:31])

    headers = ["名称", "品牌", "规格", "数量", "单位", "单价(¥)", "小计(¥)", "购买节点", "状态"]
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="3B332A", end_color="3B332A", fill_type="solid")
    money_fmt = "#,##0.00"
    widths = [28, 14, 22, 8, 8, 12, 12, 12, 10]

    for cat in cats:
        ws = wb[cat[:31]]
        ws.merge_cells("A1:I1")
        ws["A1"] = (
            f"{studio.get('name', 'Studio OS')} · {project.get('name', '')} · 物料清单（{cat}）"
        )
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 26

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        r = 3
        cat_total = 0.0
        for m in materials:
            if (m.get("category") or "其他") != cat:
                continue
            qty = float(m.get("quantity") or 0)
            price = float(m.get("unit_price") or 0)
            line = round(qty * price, 2)
            cat_total += line
            vals = [
                m.get("name"), m.get("brand"), m.get("spec"), qty, m.get("unit"),
                price, line, m.get("purchase_stage"), m.get("status"),
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.border = border
                if col in (4, 5, 6, 7):
                    cell.number_format = money_fmt
                    cell.alignment = Alignment(horizontal="right")
            r += 1
        ws.cell(row=r, column=1, value="本类合计")
        tcell = ws.cell(row=r, column=7, value=round(cat_total, 2))
        tcell.number_format = money_fmt
        tcell.font = Font(bold=True)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_bom_pdf(materials: list, project: dict, studio: dict) -> bytes:
    import os

    from fpdf import FPDF

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    font_path = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"
    if os.path.exists(font_path):
        pdf.add_font("UF", "", font_path, uni=True)
        ff = "UF"
    else:
        ff = "Helvetica"

    pdf.set_font(ff, size=15)
    pdf.cell(0, 9, studio.get("name", "Studio OS") or "Studio OS", ln=True)
    pdf.set_font(ff, size=11)
    pdf.cell(0, 7, f"物料清单 (BOM) · {project.get('name', '')} · {project.get('code', '')}", ln=True)
    pdf.ln(3)

    cols = [
        ("名称", 46), ("品牌", 24), ("规格", 34),
        ("数量", 16), ("单价", 22), ("小计", 24), ("状态", 18),
    ]

    def _header():
        pdf.set_font(ff, "", 9)
        pdf.set_fill_color(59, 51, 42)
        pdf.set_text_color(255, 255, 255)
        for name, w in cols:
            pdf.cell(w, 8, name, border=1, align="C", fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    _header()
    pdf.set_font(ff, "", 9)
    fill = False
    for m in materials:
        if fill:
            pdf.set_fill_color(240, 235, 227)
        else:
            pdf.set_fill_color(255, 255, 255)
        qty = float(m.get("quantity") or 0)
        price = float(m.get("unit_price") or 0)
        line = qty * price
        qty_str = f"{qty:.0f}" if qty == int(qty) else f"{qty:g}"
        vals = [
            str(m.get("name", "")), str(m.get("brand", "")), str(m.get("spec", "")),
            qty_str, f"¥{price:,.0f}", f"¥{line:,.0f}", str(m.get("status", "")),
        ]
        for (name, w), val in zip(cols, vals):
            align = "L" if name in ("名称", "品牌", "规格") else "C"
            pdf.cell(w, 7, val, border=1, align=align, fill=fill)
        pdf.ln()
        fill = not fill

    pdf.ln(2)
    total = sum(
        float(m.get("quantity") or 0) * float(m.get("unit_price") or 0) for m in materials
    )
    pdf.set_font(ff, "", 10)
    pdf.cell(0, 8, f"物料总估值：¥{total:,.2f}", ln=True)
    return pdf.output()
