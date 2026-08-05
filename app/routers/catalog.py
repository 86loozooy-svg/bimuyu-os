"""统一单价目录 + 空间预设 路由."""

import csv
import io
import json

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates

from app.auth import get_current_user, require_admin
from app.config import BASE_DIR
from app.database import db_session, row_to_dict, rows_to_list

router = APIRouter(prefix="/app/catalog", tags=["catalog"])
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

CATEGORIES = ["水电", "木工", "瓦工", "油工", "其他"]
CATEGORY_ORDER = {c: i for i, c in enumerate(CATEGORIES)}


def _parse_csv_rows(raw: bytes) -> list[list[str]]:
    """Parse CSV/TSV bytes into list of rows."""
    content = raw.decode("utf-8-sig")
    sample = content[:2048]
    delimiter = "\t" if "\t" in sample and "," not in sample else ","
    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    return [row for row in reader]


def _parse_excel_rows(raw: bytes) -> list[list[str]]:
    """Parse xlsx bytes into list of string rows."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# 页面
# ---------------------------------------------------------------------------
@router.get("", response_class=HTMLResponse)
async def catalog_index(
    request: Request,
    tab: str = "catalog",
    user: dict = Depends(get_current_user),
):
    with db_session() as conn:
        catalog_items = rows_to_list(
            conn.execute(
                "SELECT * FROM price_catalog ORDER BY "
                "CASE category WHEN '水电' THEN 1 WHEN '木工' THEN 2 "
                "WHEN '瓦工' THEN 3 WHEN '油工' THEN 4 ELSE 5 END, "
                "sort_order, id"
            ).fetchall()
        )
        presets = rows_to_list(
            conn.execute("SELECT * FROM space_presets ORDER BY id").fetchall()
        )
        for p in presets:
            p["parsed_items"] = json.loads(p["items_json"])

    return templates.TemplateResponse(
        "app/catalog/index.html",
        {
            "request": request,
            "user": user,
            "tab": tab,
            "catalog_items": catalog_items,
            "presets": presets,
            "categories": CATEGORIES,
        },
    )


# ---------------------------------------------------------------------------
# 单价目录 CRUD
# ---------------------------------------------------------------------------
@router.post("/add")
async def catalog_add(
    user: dict = Depends(require_admin),
    name: str = Form(...),
    unit: str = Form("㎡"),
    price: float = Form(0),
    category: str = Form("其他"),
    note: str = Form(""),
    factor: float | None = Form(None),
    sort_order: int = Form(0),
):
    with db_session() as conn:
        conn.execute(
            """
            INSERT INTO price_catalog (name, unit, price, category, note, factor, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (name, unit, price, category, note or None, factor, sort_order),
        )
    return RedirectResponse(url="/app/catalog?tab=catalog", status_code=303)


@router.post("/{item_id}/update")
async def catalog_update(
    item_id: int,
    user: dict = Depends(require_admin),
    name: str = Form(...),
    unit: str = Form("㎡"),
    price: float = Form(0),
    category: str = Form("其他"),
    note: str = Form(""),
    factor: float | None = Form(None),
    sort_order: int = Form(0),
):
    with db_session() as conn:
        conn.execute(
            """
            UPDATE price_catalog SET name=?, unit=?, price=?, category=?, note=?, factor=?, sort_order=?
            WHERE id=?
            """,
            (name, unit, price, category, note or None, factor, sort_order, item_id),
        )
    return RedirectResponse(url="/app/catalog?tab=catalog", status_code=303)


@router.post("/{item_id}/delete")
async def catalog_delete(item_id: int, user: dict = Depends(require_admin)):
    with db_session() as conn:
        conn.execute("DELETE FROM price_catalog WHERE id=?", (item_id,))
    return RedirectResponse(url="/app/catalog?tab=catalog", status_code=303)


# ---------------------------------------------------------------------------
# CSV / Excel 导入 / CSV 导出
# ---------------------------------------------------------------------------
@router.post("/import")
async def catalog_import(
    user: dict = Depends(require_admin),
    file: UploadFile = File(...),
):
    raw = await file.read()
    filename = file.filename or ""

    # 根据文件类型选择解析方式
    if filename.endswith((".xlsx", ".xls")):
        rows = _parse_excel_rows(raw)
    else:
        rows = _parse_csv_rows(raw)

    if not rows:
        return RedirectResponse(url="/app/catalog?tab=catalog", status_code=303)

    # find header row
    header_row = rows[0]
    col_map = {}
    for idx, h in enumerate(header_row):
        h_lower = h.strip().lower()
        if h_lower in ("name", "名称", "项目名称"):
            col_map["name"] = idx
        elif h_lower in ("unit", "单位"):
            col_map["unit"] = idx
        elif h_lower in ("price", "单价", "价格"):
            col_map["price"] = idx
        elif h_lower in ("category", "分类", "工种"):
            col_map["category"] = idx
        elif h_lower in ("note", "备注", "说明"):
            col_map["note"] = idx
        elif h_lower in ("factor", "系数"):
            col_map["factor"] = idx

    data_rows = rows[1:] if col_map else rows
    if not col_map:
        # no header — try positional: name, unit, price, category, note
        col_map = {"name": 0, "unit": 1, "price": 2, "category": 3, "note": 4}

    inserted = 0
    with db_session() as conn:
        for row in data_rows:
            if not row or all(c.strip() == "" for c in row):
                continue
            try:
                name = row[col_map["name"]].strip() if len(row) > col_map.get("name", 0) else ""
                if not name:
                    continue
                unit = row[col_map["unit"]].strip() if "unit" in col_map and len(row) > col_map["unit"] else "㎡"
                price_str = row[col_map["price"]].strip() if "price" in col_map and len(row) > col_map["price"] else "0"
                price = float(price_str) if price_str else 0.0
                category = row[col_map["category"]].strip() if "category" in col_map and len(row) > col_map["category"] else "其他"
                if category not in CATEGORIES:
                    category = "其他"
                note = row[col_map["note"]].strip() if "note" in col_map and len(row) > col_map["note"] else None
                factor_str = row[col_map["factor"]].strip() if "factor" in col_map and len(row) > col_map["factor"] else ""
                factor = float(factor_str) if factor_str else None

                conn.execute(
                    """
                    INSERT INTO price_catalog (name, unit, price, category, note, factor, sort_order)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (name, unit, price, category, note or None, factor, inserted + 1),
                )
                inserted += 1
            except (ValueError, IndexError):
                continue

    return RedirectResponse(url=f"/app/catalog?tab=catalog&imported={inserted}", status_code=303)


@router.get("/export")
async def catalog_export(user: dict = Depends(get_current_user)):
    with db_session() as conn:
        items = rows_to_list(
            conn.execute(
                "SELECT * FROM price_catalog ORDER BY "
                "CASE category WHEN '水电' THEN 1 WHEN '木工' THEN 2 "
                "WHEN '瓦工' THEN 3 WHEN '油工' THEN 4 ELSE 5 END, "
                "sort_order, id"
            ).fetchall()
        )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["名称", "单位", "单价", "分类", "系数", "备注"])
    for it in items:
        writer.writerow([
            it["name"],
            it["unit"],
            it["price"],
            it["category"],
            it["factor"] or "",
            it["note"] or "",
        ])

    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="price_catalog.csv"'},
    )


# ---------------------------------------------------------------------------
# 空间预设 CRUD
# ---------------------------------------------------------------------------
@router.post("/presets/add")
async def preset_add(
    user: dict = Depends(require_admin),
    name: str = Form(...),
    space_type: str = Form(""),
    items_json: str = Form("[]"),
):
    # validate JSON
    try:
        parsed = json.loads(items_json)
        if not isinstance(parsed, list):
            raise ValueError
    except (json.JSONDecodeError, ValueError):
        raise HTTPException(status_code=400, detail="无效的 JSON")

    with db_session() as conn:
        conn.execute(
            "INSERT INTO space_presets (name, space_type, items_json) VALUES (?, ?, ?)",
            (name, space_type or name, items_json),
        )
    return RedirectResponse(url="/app/catalog?tab=presets", status_code=303)


@router.post("/presets/{preset_id}/delete")
async def preset_delete(preset_id: int, user: dict = Depends(require_admin)):
    with db_session() as conn:
        conn.execute("DELETE FROM space_presets WHERE id=?", (preset_id,))
    return RedirectResponse(url="/app/catalog?tab=presets", status_code=303)


# ---------------------------------------------------------------------------
# JSON API — 供报价编辑器前端调用
# ---------------------------------------------------------------------------
@router.get("/api/items")
async def api_items(user: dict = Depends(get_current_user)):
    with db_session() as conn:
        items = rows_to_list(
            conn.execute("SELECT * FROM price_catalog ORDER BY category, sort_order, id").fetchall()
        )
    # group by category
    grouped = {}
    for it in items:
        cat = it["category"]
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(it)
    return {"categories": CATEGORIES, "items": grouped}


@router.get("/api/presets")
async def api_presets(user: dict = Depends(get_current_user)):
    with db_session() as conn:
        presets = rows_to_list(
            conn.execute("SELECT * FROM space_presets ORDER BY id").fetchall()
        )
        for p in presets:
            p["items"] = json.loads(p.pop("items_json"))
    return {"presets": presets}
